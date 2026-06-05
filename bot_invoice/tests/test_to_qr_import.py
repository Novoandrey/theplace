"""Тесты сборки файла импорта Quick Resto (to_qr_import).

Бизнес-логика: рубли из копеек, цена «с НДС» ведущая, формат ставки НДС, сборка строк,
нерешённые строки, сходимость итогов; round-trip CSV/XLSX.
"""
import csv

import pytest

import to_qr_import as tqi


def _item(n=1, no=20000, rate=10, vat=2000, wth=22000, qr=True, qty=2.0, art="X1", note=None):
    it = {"n": n, "sum_no_vat_kop": no, "vat_rate": rate, "vat_sum_kop": vat,
          "sum_with_vat_kop": wth, "name": f"товар {n}"}
    if qr:
        it["qr"] = {"art": art, "name": f"qr {art}", "type": "Ингредиент",
                    "unit": "кг", "qty": qty, "note": note}
    return it


def _doc(items):
    totals = {"sum_no_vat_kop": sum(i["sum_no_vat_kop"] for i in items),
              "vat_sum_kop": sum(i["vat_sum_kop"] for i in items),
              "sum_with_vat_kop": sum(i["sum_with_vat_kop"] for i in items)}
    return {"document": {"number": "T-1"}, "items": items, "totals": totals}


def test_rub_kopecks_to_rubles():
    assert tqi.rub(100) == 1.0
    assert tqi.rub(150) == 1.5
    assert tqi.rub(1) == 0.01


def test_build_ek2029_rows_and_totals(ek2029):
    rows, unresolved, _flags, recon = tqi.build(ek2029)
    assert len(rows) == 11
    assert unresolved == []
    for _label, (got, exp) in recon.items():
        assert got == exp
    assert recon["без НДС"][0] == 1477280
    assert recon["НДС"][0] == 187933
    assert recon["с НДС"][0] == 1665213


def test_unresolved_item_excluded():
    rows, unresolved, _, _ = tqi.build(_doc([_item(n=1), _item(n=2, qr=False)]))
    assert unresolved == [2]
    assert len(rows) == 1
    assert rows[0][0] == 1   # № строки


def test_vat_rate_string_format():
    rows, *_ = tqi.build(_doc([_item(rate=22)]))
    assert rows[0][8] == "22.0 %"   # колонка «НДС»


def test_price_is_with_vat_per_unit():
    # сумма с НДС 22000 коп = 220.00 руб, qty 2 → цена с НДС 110.00
    rows, *_ = tqi.build(_doc([_item(wth=22000, qty=2.0)]))
    assert rows[0][6] == pytest.approx(110.00)   # колонка «Цена с НДС»


def test_note_becomes_flag():
    _, _, flags, _ = tqi.build(_doc([_item(note="проверить вес")]))
    assert flags and "проверить вес" in flags[0]


def test_write_files_roundtrip(ek2029, tmp_path):
    rows, *_ = tqi.build(ek2029)
    out = str(tmp_path / "imp")
    tqi.write_files(rows, out)

    with open(out + ".csv", encoding="utf-8-sig", newline="") as f:
        data = list(csv.reader(f, delimiter="\t"))
    assert data[0] == tqi.COLS
    assert len(data) == 12   # заголовок + 11 строк

    from openpyxl import load_workbook
    ws = load_workbook(out + ".xlsx").active
    body = [[c.value for c in r] for r in ws.iter_rows()][1:]
    assert len(body) == 11
    assert round(sum(r[10] for r in body), 2) == 16652.13   # сумма «с НДС»
